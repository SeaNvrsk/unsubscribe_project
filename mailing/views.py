from django.shortcuts import render, redirect
from .forms import UploadFileForm
import openpyxl
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
import csv
from django.http import HttpResponse
from django.shortcuts import render
from django.utils.timezone import now
from django.core.mail import send_mail
import uuid
import logging
from .models import Subscriber
from io import BytesIO
from django.http import HttpResponse
import openpyxl
from .models import Subscriber

logger = logging.getLogger('unsubscribe')           # для действий пользователя
error_logger = logging.getLogger('django')          # для ошибок (в т.ч. отправка на почту)

def unsubscribe_view(request):
    token = request.GET.get("token")

    if token:
        try:
            uuid_obj = uuid.UUID(token)
        except ValueError:
            logger.warning(f"Token inválido recibido: {token} | IP: {request.META.get('REMOTE_ADDR')}")
            return render(request, "mailing/invalid.html")

        try:
            subscriber = Subscriber.objects.get(unsubscribe_token=uuid_obj)
            if not subscriber.is_unsubscribed:
                subscriber.is_unsubscribed = True
                subscriber.save()

                logger.info(f"El usuario {subscriber.email} se dio de baja correctamente. IP: {request.META.get('REMOTE_ADDR')}")

                try:
                    logger.info("📧 Enviando notificación de baja por email...")
                    send_mail(
                        subject="Usuario se dio de baja",
                        message=f"El usuario {subscriber.email} se dio de baja del boletín el {now().strftime('%d/%m/%Y %H:%M')}.",
                        from_email="notificaciones@recovia.solutions",
                        recipient_list=["reclamaciones@recovia.mx"],
                        fail_silently=False,
                    )
                    logger.info("📧 Email de notificación enviado exitosamente.")
                except Exception as e:
                    error_logger.error(f"❌ Fallo al enviar email de baja: {e}", exc_info=True)

            return render(request, "mailing/success.html", context={"email": subscriber.email})

        except Subscriber.DoesNotExist:
            logger.warning(f"Token válido pero no encontrado: {token} | IP: {request.META.get('REMOTE_ADDR')}")
            return render(request, "mailing/invalid.html")

        except Exception as e:
            error_logger.error(f"❌ Error inesperado en la vista de baja: {str(e)}", exc_info=True)
            return render(request, "mailing/error.html")

    logger.warning(f"Solicitud sin token recibida. IP: {request.META.get('REMOTE_ADDR')}")
    return render(request, "mailing/invalid.html")

@login_required
def upload_excel(request):
    logger.info(">>> upload_excel called")
    if request.method == "POST":
        logger.info("POST request received")

        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            logger.info("Form is valid")

            excel_file = request.FILES["file"]
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active

            count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                email = row[0]
                logger.info(f"Read row: {row}")
                if email:
                    _, created = Subscriber.objects.get_or_create(email=email)
                    if created:
                        count += 1

            logger.info(f"Imported {count} new emails")
            return render(request, "mailing/upload_success.html", {"count": count})
        else:
            logger.warning("Form is invalid")
    else:
        form = UploadFileForm()
        logger.info("GET request received")

    return render(request, "index.html", {"form": form})


@login_required
def export_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="subscribers.csv"'

    writer = csv.writer(response)
    writer.writerow(['email', 'unsubscribe_url'])

    for sub in Subscriber.objects.filter(is_unsubscribed=False):
        unsubscribe_url = f"https://unsubscribe.recovia.solutions/unsubscribe/?token={sub.unsubscribe_token}"
        writer.writerow([sub.email, unsubscribe_url])

    return response

def index(request):
    if request.method == "POST":
        print("File upload received")
        print(request.FILES)  # Добавь временно
        # дальше логика загрузки
    return render(request, "index.html")

def upload_subscribers(request):
    if request.method == 'POST' and request.FILES['file']:
        wb = openpyxl.load_workbook(request.FILES['file'])
        sheet = wb.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            email = row[0]
            if email:
                Subscriber.objects.get_or_create(
                    email=email,
                    defaults={'unsubscribe_token': uuid.uuid4()})
    return redirect('index')


def export_subscribers(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dado de baja"
    ws.append(['Email', 'Fecha de baja', 'Token'])

    unsubscribed = Subscriber.objects.filter(is_unsubscribed=True)

    for s in unsubscribed:
        ws.append([
            s.email,
            s.created_at.strftime('%Y-%m-%d %H:%M'),
            str(s.unsubscribe_token)
        ])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=dados_de_baja.xlsx'
    return response


@login_required
def export_all_subscribers(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Email', 'Dado de baja', 'Token'])

    for s in Subscriber.objects.all():
        ws.append([s.email, 'Sí' if s.is_unsubscribed else 'No', str(s.unsubscribe_token)])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=subscribers_all.xlsx'
    wb.save(response)
    return response