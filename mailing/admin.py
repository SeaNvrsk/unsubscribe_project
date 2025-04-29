from django.contrib import admin
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from .models import Subscriber

@admin.action(description="Exportar seleccionados a Excel")
def export_to_excel(modeladmin, request, queryset):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Подписчики"

    # Заголовки
    ws.append(['Email', 'Отписан?', 'Unsubscribe URL'])

    for obj in queryset:
        ws.append([
            obj.email,
            'Да' if obj.is_unsubscribed else 'Нет',
            f"https://unsubscribe.recovia.solutions/unsubscribe/?token={obj.unsubscribe_token}"
        ])

    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=subscribers.xlsx'
    wb.save(response)
    return response

@admin.register(Subscriber)
class SubscriberAdmin(admin.ModelAdmin):
    list_display = ('email', 'is_unsubscribed', 'unsubscribe_token', 'created_at')
    list_filter = ('is_unsubscribed',)
    search_fields = ('email',)
    ordering = ('email',)
    change_list_template = "admin/mailing/subscriber/change_list.html"
    actions = [export_to_excel]

    def changelist_view(self, request, extra_context=None):
        count_all = Subscriber.objects.count()
        count_active = Subscriber.objects.filter(is_unsubscribed=False).count()
        count_unsubscribed = Subscriber.objects.filter(is_unsubscribed=True).count()

        extra_context = extra_context or {}
        extra_context['count_all'] = count_all
        extra_context['count_active'] = count_active
        extra_context['count_unsubscribed'] = count_unsubscribed
        return super().changelist_view(request, extra_context=extra_context)